# -*- coding: utf-8 -*-
# @time:2019/4/29 13:27
# Author:殇殇
# @file:do_excel.py
# @fuction: 操作Excel的读与写
from openpyxl import load_workbook

class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.check_sql = None


class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        cases=[]
        for i in range(2,sheet.max_row+1):
            row_cases = Case() # 实例化对象
            row_cases.case_id = sheet.cell(row=i,column=1).value
            row_cases.title = sheet.cell (row=i, column=2).value
            row_cases.url = sheet.cell (row=i, column=3).value
            row_cases.data = sheet.cell (row=i, column=4).value
            row_cases.method = sheet.cell (row=i, column=5).value
            row_cases.expected = sheet.cell(row=i,column=6).value
            row_cases.check_sql = sheet.cell(row=i,column=9).value
            cases.append(row_cases)
        wb.close()
        return cases

    def write_back(self,row,actual,result):
        wb= load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,7).value = actual
        sheet.cell (row, 8).value = result
        wb.save(self.file_name)
        wb.close()






